"""
Rent Comp Analysis Builder
--------------------------
Reads Yardi "Property Detail - Rental Rate History" files from ./comps/
and generates a formatted Excel workbook with 4 sheets:
  1) Summary  2) Rent Trend by Bedroom  3) PSF Trend by Bedroom  4) Unit Mix Comparison

Usage: Place Yardi .xlsx/.xlsm files in ./comps/, edit SUBJECT below, run:
  python rent_comp_builder.py
"""

import os
import sys
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════
# SUBJECT PROPERTY — edit these values
# ══════════════════════════════════════════════════════════════════════
SUBJECT = {
    "name": "Subject Property",
    "address": "123 Main Street",
    "year_built": 2024,
    "unit_mix": [
        # {"type": "Studio",  "units": 20, "avg_sf": 500, "rent": 1400},
        {"type": "1BR", "units": 100, "avg_sf": 750, "rent": 1800},
        {"type": "2BR", "units": 80,  "avg_sf": 1050, "rent": 2400},
        {"type": "3BR", "units": 20,  "avg_sf": 1300, "rent": 3200},
    ],
}

COMPS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "comps")
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Rent_Comp_Analysis.xlsx")

# ══════════════════════════════════════════════════════════════════════
# Colors / styles
# ══════════════════════════════════════════════════════════════════════
NAVY = "1B2A4A"
STEEL = "4A6FA5"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
SUBJECT_BLUE = "DCE6F1"

HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(name="Arial", bold=True, color=WHITE, size=10)
SUBJ_FILL = PatternFill("solid", fgColor=SUBJECT_BLUE)
ALT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FONT = Font(name="Arial", size=12, bold=True, color=NAVY)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

CHART_COLORS = [
    "4A6FA5", "C9A84C", "2E7D4F", "B94A48", "8A7A9A",
    "6A9A96", "A8856A", "A8706A", "6A9A7A", "9A7A7A",
]

BED_TYPES = ["Studio", "1BR", "2BR", "3BR"]
BED_ORDER = {t: i for i, t in enumerate(BED_TYPES)}


# ══════════════════════════════════════════════════════════════════════
# Yardi parser
# ══════════════════════════════════════════════════════════════════════
def classify_unit_type(raw):
    s = raw.strip().lower()
    if "two bedroom" in s or "2 bed" in s or "2 br" in s or "2br" in s:
        return "2BR"
    if "three bedroom" in s or "3 bed" in s or "3 br" in s or "3br" in s:
        return "3BR"
    if "one bedroom" in s or "1 bed" in s or "1 br" in s or "1br" in s:
        return "1BR"
    if "studio" in s or "alcove" in s or "efficiency" in s:
        return "Studio"
    return None


def safe_float(cell):
    if cell.value is None:
        return None
    try:
        return float(cell.value)
    except (ValueError, TypeError):
        return None


def parse_yardi_file(path):
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARNING: Cannot open {os.path.basename(path)}: {e}")
        return None

    ws1 = wb.worksheets[0]
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else None

    rows = list(ws1.iter_rows(min_row=1, max_col=15, values_only=False))
    if len(rows) < 14:
        print(f"  WARNING: Too few rows in {os.path.basename(path)}, skipping")
        wb.close()
        return None

    def cv(r, c):
        try:
            return rows[r][c].value
        except (IndexError, AttributeError):
            return None

    prop_name = str(cv(2, 1) or "").strip()
    address = str(cv(3, 1) or "").strip()
    market = str(cv(4, 1) or "").strip()
    submarket = str(cv(5, 1) or "").strip()
    total_units = 0
    try:
        total_units = int(float(cv(6, 1) or 0))
    except (ValueError, TypeError):
        pass

    if not prop_name:
        print(f"  WARNING: No property name in {os.path.basename(path)}, skipping")
        wb.close()
        return None

    header_row_idx = 12
    for i in range(9, 20):
        if i < len(rows):
            val = str(cv(i, 0) or "").strip().lower()
            if val == "unit type":
                header_row_idx = i
                break

    months = []
    for c in range(3, 15):
        val = cv(header_row_idx, c)
        if val:
            months.append(str(val).strip())

    psf_rows = None
    if ws2:
        psf_rows = list(ws2.iter_rows(min_row=1, max_col=15, values_only=False))

    floorplans = []
    for i in range(header_row_idx + 1, len(rows)):
        unit_type_raw = str(cv(i, 0) or "").strip()
        units_val = safe_float(rows[i][1])
        sqft_val = safe_float(rows[i][2])

        if not unit_type_raw and units_val and units_val > 0:
            continue
        if not unit_type_raw:
            continue

        units = int(units_val) if units_val else 0
        sqft = int(sqft_val) if sqft_val else 0
        if units == 0:
            continue

        bed_type = classify_unit_type(unit_type_raw)
        if bed_type is None:
            continue

        monthly_rents = {}
        for mi, m in enumerate(months):
            v = safe_float(rows[i][3 + mi])
            if v and v > 0:
                monthly_rents[m] = v

        monthly_psf = {}
        if psf_rows and i < len(psf_rows):
            for mi, m in enumerate(months):
                try:
                    v = safe_float(psf_rows[i][3 + mi])
                    if v and v > 0:
                        monthly_psf[m] = v
                except (IndexError, AttributeError):
                    pass

        floorplans.append({
            "raw_type": unit_type_raw,
            "bed_type": bed_type,
            "units": units,
            "sqft": sqft,
            "monthly_rents": monthly_rents,
            "monthly_psf": monthly_psf,
        })

    wb.close()

    if not floorplans:
        print(f"  WARNING: No valid floorplans in {os.path.basename(path)}, skipping")
        return None

    return {
        "name": prop_name,
        "address": address,
        "market": market,
        "submarket": submarket,
        "total_units": total_units,
        "months": months,
        "floorplans": floorplans,
    }


# ══════════════════════════════════════════════════════════════════════
# Aggregation helpers
# ══════════════════════════════════════════════════════════════════════
def wtd_avg(floorplans, value_key):
    total_units = sum(fp["units"] for fp in floorplans)
    if total_units == 0:
        return 0
    return sum(fp[value_key] * fp["units"] for fp in floorplans) / total_units


def get_bed_floorplans(prop, bed_type):
    return [fp for fp in prop["floorplans"] if fp["bed_type"] == bed_type]


def get_wtd_rent_for_month(floorplans, month):
    fps_with = [(fp, fp["monthly_rents"].get(month)) for fp in floorplans]
    fps_with = [(fp, v) for fp, v in fps_with if v is not None]
    if not fps_with:
        return None
    total_u = sum(fp["units"] for fp, _ in fps_with)
    if total_u == 0:
        return None
    return sum(v * fp["units"] for fp, v in fps_with) / total_u


def get_wtd_psf_for_month(floorplans, month):
    fps_with = [(fp, fp["monthly_psf"].get(month)) for fp in floorplans]
    fps_with = [(fp, v) for fp, v in fps_with if v is not None]
    if not fps_with:
        return None
    total_u = sum(fp["units"] for fp, _ in fps_with)
    if total_u == 0:
        return None
    return sum(v * fp["units"] for fp, v in fps_with) / total_u


def build_subject_prop(subj_dict):
    floorplans = []
    for um in subj_dict["unit_mix"]:
        bed = um["type"]
        if bed not in BED_ORDER:
            bed = classify_unit_type(bed) or bed
        floorplans.append({
            "raw_type": um["type"],
            "bed_type": bed,
            "units": um["units"],
            "sqft": um["avg_sf"],
            "monthly_rents": {},
            "monthly_psf": {},
            "rent": um["rent"],
        })
    total = sum(um["units"] for um in subj_dict["unit_mix"])
    return {
        "name": subj_dict["name"],
        "address": subj_dict.get("address", ""),
        "market": "",
        "submarket": "",
        "total_units": total,
        "year_built": subj_dict.get("year_built", ""),
        "months": [],
        "floorplans": floorplans,
        "is_subject": True,
    }


# ══════════════════════════════════════════════════════════════════════
# Style helper
# ══════════════════════════════════════════════════════════════════════
def style_cell(ws, row, col, value, font=None, fill=None, fmt=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def write_header_row(ws, row, headers, col_start=1):
    for ci, h in enumerate(headers):
        style_cell(ws, row, col_start + ci, h,
                   font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center", wrap_text=True),
                   border=THIN_BORDER)


# ══════════════════════════════════════════════════════════════════════
# Sheet 1: Summary
# ══════════════════════════════════════════════════════════════════════
def build_summary_sheet(wb, subject, comps):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = NAVY

    all_props = [subject] + comps
    months = []
    for p in comps:
        if p["months"]:
            months = p["months"]
            break

    style_cell(ws, 1, 1, "Rent Comp Analysis", font=Font(name="Arial", size=14, bold=True, color=NAVY))
    style_cell(ws, 2, 1, f"Subject: {subject['name']}", font=Font(name="Arial", size=10, color="666666"))
    style_cell(ws, 3, 1, f"Generated: {__import__('datetime').date.today().strftime('%B %d, %Y')}", font=Font(name="Arial", size=10, color="666666"))

    headers = ["#", "Property", "Address", "Year Built", "Units", "Avg SF",
               "Most Recent Avg Rent", "Most Recent $/SF"]
    row = 5
    write_header_row(ws, row, headers)

    for pi, prop in enumerate(all_props):
        row += 1
        is_subj = prop.get("is_subject", False)
        fps = prop["floorplans"]

        total_u = prop["total_units"] or sum(fp["units"] for fp in fps)
        avg_sf = wtd_avg(fps, "sqft") if fps else 0

        if is_subj:
            avg_rent = wtd_avg([{**fp, "sqft": fp.get("rent", 0)} for fp in fps], "sqft") if fps else 0
            avg_rent = sum(fp.get("rent", 0) * fp["units"] for fp in fps) / max(total_u, 1)
            avg_psf = avg_rent / avg_sf if avg_sf > 0 else 0
        else:
            most_recent = months[-1] if months else None
            if most_recent:
                fps_with = [(fp, fp["monthly_rents"].get(most_recent)) for fp in fps]
                fps_with = [(fp, v) for fp, v in fps_with if v is not None]
                total_with = sum(fp["units"] for fp, _ in fps_with)
                avg_rent = sum(v * fp["units"] for fp, v in fps_with) / max(total_with, 1) if fps_with else 0
            else:
                avg_rent = 0
            avg_psf = avg_rent / avg_sf if avg_sf > 0 else 0

        row_fill = SUBJ_FILL if is_subj else (ALT_FILL if (pi % 2 == 1) else None)
        row_font = BOLD_FONT if is_subj else BODY_FONT

        style_cell(ws, row, 1, pi + 1 if not is_subj else "S", font=row_font, fill=row_fill, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, prop["name"] + (" (Subject)" if is_subj else ""), font=row_font, fill=row_fill, border=THIN_BORDER)
        style_cell(ws, row, 3, prop.get("address", ""), font=row_font, fill=row_fill, border=THIN_BORDER)
        style_cell(ws, row, 4, prop.get("year_built", ""), font=row_font, fill=row_fill, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 5, total_u, font=row_font, fill=row_fill, border=THIN_BORDER, fmt="#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 6, round(avg_sf), font=row_font, fill=row_fill, border=THIN_BORDER, fmt="#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 7, round(avg_rent), font=row_font, fill=row_fill, border=THIN_BORDER, fmt="$#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 8, round(avg_psf, 2), font=row_font, fill=row_fill, border=THIN_BORDER, fmt="$#,##0.00", alignment=Alignment(horizontal="center"))

    # Comp Averages row
    if comps:
        row += 1
        avg_fill = PatternFill("solid", fgColor="E8E0D0")
        avg_font = Font(name="Arial", size=10, bold=True)
        comp_total_u = [c["total_units"] or sum(fp["units"] for fp in c["floorplans"]) for c in comps]
        comp_avg_sf = [wtd_avg(c["floorplans"], "sqft") for c in comps]

        most_recent = months[-1] if months else None
        comp_rents = []
        for c in comps:
            if most_recent:
                fps_with = [(fp, fp["monthly_rents"].get(most_recent)) for fp in c["floorplans"]]
                fps_with = [(fp, v) for fp, v in fps_with if v is not None]
                tu = sum(fp["units"] for fp, _ in fps_with)
                r = sum(v * fp["units"] for fp, v in fps_with) / max(tu, 1) if fps_with else 0
                comp_rents.append(r)
            else:
                comp_rents.append(0)

        a_units = round(sum(comp_total_u) / len(comps))
        a_sf = round(sum(comp_avg_sf) / len(comps))
        a_rent = round(sum(comp_rents) / len(comps))
        a_psf = a_rent / a_sf if a_sf > 0 else 0

        style_cell(ws, row, 1, "", font=avg_font, fill=avg_fill, border=THIN_BORDER)
        style_cell(ws, row, 2, "COMP AVERAGE", font=avg_font, fill=avg_fill, border=THIN_BORDER)
        style_cell(ws, row, 3, "", font=avg_font, fill=avg_fill, border=THIN_BORDER)
        style_cell(ws, row, 4, "", font=avg_font, fill=avg_fill, border=THIN_BORDER)
        style_cell(ws, row, 5, a_units, font=avg_font, fill=avg_fill, border=THIN_BORDER, fmt="#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 6, a_sf, font=avg_font, fill=avg_fill, border=THIN_BORDER, fmt="#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 7, a_rent, font=avg_font, fill=avg_fill, border=THIN_BORDER, fmt="$#,##0", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 8, round(a_psf, 2), font=avg_font, fill=avg_fill, border=THIN_BORDER, fmt="$#,##0.00", alignment=Alignment(horizontal="center"))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 16


# ══════════════════════════════════════════════════════════════════════
# Sheet 2: Rent Trend by Bedroom
# ══════════════════════════════════════════════════════════════════════
def build_rent_trend_sheet(wb, subject, comps, months):
    ws = wb.create_sheet("Rent Trend by Bedroom")
    ws.sheet_properties.tabColor = STEEL

    current_row = 1
    chart_num = 0

    for bed_type in BED_TYPES:
        props_with_data = []
        for comp in comps:
            fps = get_bed_floorplans(comp, bed_type)
            if not fps:
                continue
            has_any = any(get_wtd_rent_for_month(fps, m) is not None for m in months)
            if has_any:
                props_with_data.append(comp)

        if not props_with_data:
            continue

        style_cell(ws, current_row, 1, f"{bed_type} — Weighted Avg Rent by Month", font=TITLE_FONT)
        current_row += 1

        headers = ["Property"] + months
        write_header_row(ws, current_row, headers)
        table_start = current_row
        current_row += 1

        subj_fps = get_bed_floorplans(subject, bed_type)
        subj_rent = None
        if subj_fps:
            subj_rent = sum(fp.get("rent", 0) * fp["units"] for fp in subj_fps) / max(sum(fp["units"] for fp in subj_fps), 1)

        has_subject_row = subj_rent is not None and subj_rent > 0
        if has_subject_row:
            style_cell(ws, current_row, 1, subject["name"] + " (Subject)", font=BOLD_FONT, fill=SUBJ_FILL, border=THIN_BORDER)
            for mi, m in enumerate(months):
                style_cell(ws, current_row, 2 + mi, round(subj_rent), font=BOLD_FONT, fill=SUBJ_FILL, border=THIN_BORDER, fmt="$#,##0", alignment=Alignment(horizontal="center"))
            current_row += 1

        for ci, comp in enumerate(props_with_data):
            fps = get_bed_floorplans(comp, bed_type)
            row_fill = ALT_FILL if (ci % 2 == 0) else None
            style_cell(ws, current_row, 1, comp["name"], font=BODY_FONT, fill=row_fill, border=THIN_BORDER)
            for mi, m in enumerate(months):
                val = get_wtd_rent_for_month(fps, m)
                style_cell(ws, current_row, 2 + mi, round(val) if val else None, font=BODY_FONT, fill=row_fill, border=THIN_BORDER, fmt="$#,##0", alignment=Alignment(horizontal="center"))
            current_row += 1

        table_end = current_row - 1
        num_data_rows = table_end - table_start

        chart = LineChart()
        chart.title = f"{bed_type} — Monthly Rent Trend"
        chart.y_axis.title = "Rent ($)"
        chart.x_axis.title = "Month"
        chart.y_axis.auto = True
        chart.y_axis.delete = False
        chart.y_axis.crossesAt = 0
        chart.y_axis.scaling.min = None
        chart.width = 28
        chart.height = 14
        chart.style = 10

        cats = Reference(ws, min_col=2, max_col=1 + len(months), min_row=table_start)

        color_idx = 0
        for dr in range(1, num_data_rows + 1):
            data_ref = Reference(ws, min_col=2, max_col=1 + len(months), min_row=table_start + dr)
            label_ref = Reference(ws, min_col=1, min_row=table_start + dr)
            chart.add_data(data_ref, from_rows=True)
            s = chart.series[-1]
            s.title = SeriesLabel(label_ref)

            if has_subject_row and dr == 1:
                s.graphicalProperties.line.dashStyle = "dash"
                s.graphicalProperties.line.solidFill = STEEL
                s.graphicalProperties.line.width = 28000
            else:
                c = CHART_COLORS[color_idx % len(CHART_COLORS)]
                s.graphicalProperties.line.solidFill = c
                s.graphicalProperties.line.width = 22000
                color_idx += 1

            s.marker.symbol = "circle"
            s.marker.size = 5

        chart.set_categories(cats)
        chart.legend.position = "b"

        min_vals = []
        for dr in range(1, num_data_rows + 1):
            for mc in range(2, 2 + len(months)):
                v = ws.cell(row=table_start + dr, column=mc).value
                if v is not None and isinstance(v, (int, float)) and v > 0:
                    min_vals.append(v)
        if min_vals:
            floor_val = min(min_vals) * 0.92
            chart.y_axis.scaling.min = round(floor_val / 50) * 50

        current_row += 1
        ws.add_chart(chart, f"A{current_row}")
        current_row += 18

    ws.column_dimensions["A"].width = 28
    for c in range(2, 2 + len(months)):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ══════════════════════════════════════════════════════════════════════
# Sheet 3: PSF Trend by Bedroom
# ══════════════════════════════════════════════════════════════════════
def build_psf_trend_sheet(wb, subject, comps, months):
    ws = wb.create_sheet("PSF Trend by Bedroom")
    ws.sheet_properties.tabColor = "C9A84C"

    current_row = 1

    for bed_type in BED_TYPES:
        props_with_data = []
        for comp in comps:
            fps = get_bed_floorplans(comp, bed_type)
            if not fps:
                continue
            has_any = any(get_wtd_psf_for_month(fps, m) is not None for m in months)
            if has_any:
                props_with_data.append(comp)

        if not props_with_data:
            continue

        style_cell(ws, current_row, 1, f"{bed_type} — Weighted Avg $/SF by Month", font=TITLE_FONT)
        current_row += 1

        headers = ["Property"] + months
        write_header_row(ws, current_row, headers)
        table_start = current_row
        current_row += 1

        subj_fps = get_bed_floorplans(subject, bed_type)
        subj_psf = None
        if subj_fps:
            subj_rent = sum(fp.get("rent", 0) * fp["units"] for fp in subj_fps) / max(sum(fp["units"] for fp in subj_fps), 1)
            subj_sf = sum(fp["sqft"] * fp["units"] for fp in subj_fps) / max(sum(fp["units"] for fp in subj_fps), 1)
            if subj_sf > 0 and subj_rent > 0:
                subj_psf = subj_rent / subj_sf

        has_subject_row = subj_psf is not None and subj_psf > 0
        if has_subject_row:
            style_cell(ws, current_row, 1, subject["name"] + " (Subject)", font=BOLD_FONT, fill=SUBJ_FILL, border=THIN_BORDER)
            for mi, m in enumerate(months):
                style_cell(ws, current_row, 2 + mi, round(subj_psf, 2), font=BOLD_FONT, fill=SUBJ_FILL, border=THIN_BORDER, fmt="$#,##0.00", alignment=Alignment(horizontal="center"))
            current_row += 1

        for ci, comp in enumerate(props_with_data):
            fps = get_bed_floorplans(comp, bed_type)
            row_fill = ALT_FILL if (ci % 2 == 0) else None
            style_cell(ws, current_row, 1, comp["name"], font=BODY_FONT, fill=row_fill, border=THIN_BORDER)
            for mi, m in enumerate(months):
                val = get_wtd_psf_for_month(fps, m)
                style_cell(ws, current_row, 2 + mi, round(val, 2) if val else None, font=BODY_FONT, fill=row_fill, border=THIN_BORDER, fmt="$#,##0.00", alignment=Alignment(horizontal="center"))
            current_row += 1

        table_end = current_row - 1
        num_data_rows = table_end - table_start

        chart = LineChart()
        chart.title = f"{bed_type} — Monthly $/SF Trend"
        chart.y_axis.title = "$/SF"
        chart.x_axis.title = "Month"
        chart.y_axis.numFmt = "$#,##0.00"
        chart.width = 28
        chart.height = 14
        chart.style = 10

        cats = Reference(ws, min_col=2, max_col=1 + len(months), min_row=table_start)

        color_idx = 0
        for dr in range(1, num_data_rows + 1):
            data_ref = Reference(ws, min_col=2, max_col=1 + len(months), min_row=table_start + dr)
            label_ref = Reference(ws, min_col=1, min_row=table_start + dr)
            chart.add_data(data_ref, from_rows=True)
            s = chart.series[-1]
            s.title = SeriesLabel(label_ref)

            if has_subject_row and dr == 1:
                s.graphicalProperties.line.dashStyle = "dash"
                s.graphicalProperties.line.solidFill = STEEL
                s.graphicalProperties.line.width = 28000
            else:
                c = CHART_COLORS[color_idx % len(CHART_COLORS)]
                s.graphicalProperties.line.solidFill = c
                s.graphicalProperties.line.width = 22000
                color_idx += 1

            s.marker.symbol = "circle"
            s.marker.size = 5

        chart.set_categories(cats)
        chart.legend.position = "b"

        min_vals = []
        for dr in range(1, num_data_rows + 1):
            for mc in range(2, 2 + len(months)):
                v = ws.cell(row=table_start + dr, column=mc).value
                if v is not None and isinstance(v, (int, float)) and v > 0:
                    min_vals.append(v)
        if min_vals:
            floor_val = min(min_vals) * 0.90
            chart.y_axis.scaling.min = round(floor_val * 4) / 4

        current_row += 1
        ws.add_chart(chart, f"A{current_row}")
        current_row += 18

    ws.column_dimensions["A"].width = 28
    for c in range(2, 2 + len(months)):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ══════════════════════════════════════════════════════════════════════
# Sheet 4: Unit Mix Comparison
# ══════════════════════════════════════════════════════════════════════
def build_unit_mix_sheet(wb, subject, comps, months):
    ws = wb.create_sheet("Unit Mix Comparison")
    ws.sheet_properties.tabColor = "2E7D4F"

    all_props = [subject] + comps

    style_cell(ws, 1, 1, "Unit Mix Comparison (% of Total Units)", font=TITLE_FONT)

    headers = ["Property", "Total Units"] + BED_TYPES
    write_header_row(ws, 3, headers)

    for pi, prop in enumerate(all_props):
        row = 4 + pi
        is_subj = prop.get("is_subject", False)
        fps = prop["floorplans"]
        total = prop["total_units"] or sum(fp["units"] for fp in fps)

        row_fill = SUBJ_FILL if is_subj else (ALT_FILL if (pi % 2 == 1) else None)
        row_font = BOLD_FONT if is_subj else BODY_FONT

        style_cell(ws, row, 1, prop["name"] + (" (Subject)" if is_subj else ""), font=row_font, fill=row_fill, border=THIN_BORDER)
        style_cell(ws, row, 2, total, font=row_font, fill=row_fill, border=THIN_BORDER, fmt="#,##0", alignment=Alignment(horizontal="center"))

        for bi, bed in enumerate(BED_TYPES):
            bed_fps = get_bed_floorplans(prop, bed)
            bed_units = sum(fp["units"] for fp in bed_fps)
            pct = bed_units / total if total > 0 else 0
            style_cell(ws, row, 3 + bi, pct, font=row_font, fill=row_fill, border=THIN_BORDER, fmt="0.0%", alignment=Alignment(horizontal="center"))

    table_end_row = 3 + len(all_props)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Unit Mix by Property"
    chart.y_axis.title = "% of Units"
    chart.y_axis.numFmt = "0%"
    chart.width = 24
    chart.height = 14
    chart.style = 10

    cats = Reference(ws, min_col=3, max_col=2 + len(BED_TYPES), min_row=3)
    chart.set_categories(cats)

    for pi in range(len(all_props)):
        data_ref = Reference(ws, min_col=3, max_col=2 + len(BED_TYPES), min_row=4 + pi)
        label_ref = Reference(ws, min_col=1, min_row=4 + pi)
        chart.add_data(data_ref, from_rows=True)
        s = chart.series[-1]
        s.title = SeriesLabel(label_ref)

        is_subj = all_props[pi].get("is_subject", False)
        c = STEEL if is_subj else CHART_COLORS[pi % len(CHART_COLORS)]
        s.graphicalProperties.solidFill = c

    chart.legend.position = "b"
    ws.add_chart(chart, f"A{table_end_row + 2}")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    for c in range(3, 3 + len(BED_TYPES)):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ══════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Rent Comp Analysis Builder")
    print("=" * 60)

    if not os.path.isdir(COMPS_DIR):
        print(f"\nERROR: Comps directory not found: {COMPS_DIR}")
        print("Create a 'comps' folder and place Yardi files inside.")
        sys.exit(1)

    files = [f for f in os.listdir(COMPS_DIR) if f.lower().endswith((".xlsx", ".xlsm"))]
    if not files:
        print(f"\nERROR: No .xlsx or .xlsm files found in {COMPS_DIR}")
        sys.exit(1)

    print(f"\nFound {len(files)} file(s) in {COMPS_DIR}")
    comps = []
    skipped = 0

    for f in sorted(files):
        path = os.path.join(COMPS_DIR, f)
        print(f"  Parsing: {f}...", end=" ")
        result = parse_yardi_file(path)
        if result:
            comps.append(result)
            total_fps = len(result["floorplans"])
            bed_types_found = sorted(set(fp["bed_type"] for fp in result["floorplans"]), key=lambda x: BED_ORDER.get(x, 99))
            print(f"OK — {result['name']} ({result['total_units']} units, {total_fps} floorplans: {', '.join(bed_types_found)})")
        else:
            skipped += 1

    if not comps:
        print("\nERROR: No files parsed successfully.")
        sys.exit(1)

    months = []
    for c in comps:
        if c["months"]:
            months = c["months"]
            break

    subject = build_subject_prop(SUBJECT)

    print(f"\nSubject: {subject['name']} ({subject['total_units']} units)")
    print(f"Comps loaded: {len(comps)}")
    if skipped:
        print(f"Skipped: {skipped}")
    print(f"Months: {months[0]} — {months[-1]}" if months else "No monthly data found")
    print(f"\nBuilding workbook...")

    wb = Workbook()
    build_summary_sheet(wb, subject, comps)
    if months:
        build_rent_trend_sheet(wb, subject, comps, months)
        build_psf_trend_sheet(wb, subject, comps, months)
    build_unit_mix_sheet(wb, subject, comps, months)

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print("=" * 60)
    print("  Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()
